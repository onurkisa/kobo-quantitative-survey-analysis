"""Reusable Excel-report writing helpers with explicit style and content inputs."""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .kobo_metadata import get_question_label


def safe_sheet_name(name: object, used_names: set[str]) -> str:
    """Return a unique Excel-compatible worksheet name and register it."""
    clean = re.sub(r"[\\/*?:\[\]]", "", str(name)).strip()[:31] or "Sheet"
    base, suffix = clean, 1
    while clean in used_names:
        marker = f"_{suffix}"
        clean = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    used_names.add(clean)
    return clean


def question_number(variable: str) -> str:
    """Return a Kobo-style numeric question prefix when available."""
    match = re.match(r"q(\d+)(?:_(\d+))?", str(variable), flags=re.I)
    if not match:
        return str(variable)
    return ".".join(part for part in match.groups() if part is not None)


def question_heading(variable: str, label: str) -> str:
    """Return a question heading without duplicating its numeric prefix."""
    number = question_number(variable)
    clean_label = str(label).strip()
    if re.match(rf"^{re.escape(number)}(?:\.|\s)", clean_label):
        return f"Question {clean_label}"
    return f"Question {number}. {clean_label}"


def write_dataframe(
    ws: Any,
    table: pd.DataFrame,
    row: int,
    style: Mapping[str, str],
    title: str | None = None,
    start_column: int = 1,
) -> int:
    """Write one standard table and return the next available row."""
    thin_grey = Side(style="thin", color=style["border_color"])
    if title:
        cell = ws.cell(row=row, column=start_column, value=title)
        cell.font = Font(bold=True, color=style["text_on_primary"])
        cell.fill = PatternFill("solid", fgColor=style["secondary_color"])
        row += 1
    values = table.copy()
    values.index.name = values.index.name or "Category"
    values = values.reset_index()
    for column, name in enumerate(values.columns, start_column):
        cell = ws.cell(row=row, column=column, value=str(name))
        cell.font = Font(bold=True, color=style["text_on_primary"])
        cell.fill = PatternFill("solid", fgColor=style["primary_color"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_grey)
    for record_index, (_, record) in enumerate(values.iterrows(), row + 1):
        is_total = str(record.iloc[0]).startswith("Total")
        for column, value in enumerate(record, start_column):
            cell = ws.cell(row=record_index, column=column, value=None if pd.isna(value) else value)
            cell.alignment = Alignment(horizontal="left" if column == start_column else "center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_grey)
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=style["total_row_color"])
            source_column = values.columns[column - start_column]
            if isinstance(value, (float, np.floating)) and (
                "Percent" in str(source_column) or str(source_column) == "Total"
            ):
                cell.number_format = "0.0"
    return row + len(values) + 1


def _clean_group_label(value: object) -> str:
    """Return a concise group label while preserving an available base n."""
    return str(value).strip()


def build_consolidated_question_matrix(
    variable: str,
    analysis_type: str,
    *,
    descriptive_outputs: Mapping[str, Mapping[Any, pd.DataFrame]],
    breakdowns: Sequence[str],
    breakdown_labels: Mapping[str, str],
) -> pd.DataFrame:
    """Combine overall and breakdown descriptives into one question matrix."""
    overall = descriptive_outputs["overall"].get(variable)
    if not isinstance(overall, pd.DataFrame) or overall.empty:
        return pd.DataFrame()

    matrix_parts: list[pd.DataFrame] = []
    if analysis_type in {"categorical_single", "open_text"}:
        retained_columns = [
            column
            for column in ["Frequency", "Percent", "Weighted_Percent"]
            if column in overall.columns
        ]
        if not retained_columns:
            return pd.DataFrame()
        valid_n = (
            int(overall["Valid_N"].max())
            if "Valid_N" in overall.columns and overall["Valid_N"].notna().any()
            else None
        )
        display_names = {
            "Frequency": "Frequency",
            "Percent": "Unweighted %",
            "Weighted_Percent": "Weighted %",
        }
        part = overall[retained_columns].rename(columns=display_names)
        if valid_n is not None:
            first_name = display_names[retained_columns[0]]
            part = part.rename(columns={first_name: f"{first_name} (base n={valid_n})"})
        part.columns = pd.MultiIndex.from_tuples([
            ("Overall", str(column)) for column in part.columns
        ])
        matrix_parts.append(part)

        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            values = table.drop(index=[
                index for index in table.index if str(index).startswith("Total")
            ], errors="ignore").drop(columns=["Total"], errors="ignore")
            if values.empty:
                continue
            transposed = values.T
            label = breakdown_labels.get(breakdown, breakdown)
            transposed.columns = pd.MultiIndex.from_tuples([
                (label, _clean_group_label(column)) for column in transposed.columns
            ])
            matrix_parts.append(transposed)

    elif analysis_type == "continuous":
        overall_values = overall.iloc[0]
        part = overall_values.to_frame(name="Overall")
        part.columns = pd.MultiIndex.from_tuples([("Overall", "All respondents")])
        matrix_parts.append(part)
        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            transposed = table.T.copy()
            label = breakdown_labels.get(breakdown, breakdown)
            transposed.columns = pd.MultiIndex.from_tuples([
                (
                    label,
                    f"{_clean_group_label(column)} (n={int(table.loc[column, 'N'])})"
                    if "N" in table.columns and pd.notna(table.loc[column, "N"])
                    else _clean_group_label(column),
                )
                for column in transposed.columns
            ])
            matrix_parts.append(transposed)

    elif analysis_type == "multiple_response":
        value_column = (
            "Weighted_Respondent_%"
            if "Weighted_Respondent_%" in overall.columns
            else "Respondent_%"
        )
        retained_columns = [
            column
            for column in ["Frequency", "Respondent_%", "Weighted_Respondent_%"]
            if column in overall.columns
        ]
        valid_n = (
            int(overall["Valid_N"].max())
            if "Valid_N" in overall.columns and overall["Valid_N"].notna().any()
            else None
        )
        display_names = {
            "Frequency": "Frequency",
            "Respondent_%": "Unweighted %",
            "Weighted_Respondent_%": "Weighted %",
        }
        part = overall.set_index("Option")[retained_columns].rename(columns=display_names)
        if valid_n is not None:
            first_name = display_names[retained_columns[0]]
            part = part.rename(columns={first_name: f"{first_name} (base n={valid_n})"})
        part.columns = pd.MultiIndex.from_tuples([
            ("Overall", str(column)) for column in part.columns
        ])
        matrix_parts.append(part)
        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            reset = table.reset_index()
            if value_column not in reset.columns:
                continue
            pivoted = reset.pivot(index="Option", columns="Group", values=value_column)
            label = breakdown_labels.get(breakdown, breakdown)
            pivoted.columns = pd.MultiIndex.from_tuples([
                (label, _clean_group_label(column)) for column in pivoted.columns
            ])
            matrix_parts.append(pivoted)

    elif analysis_type == "indicator":
        overall_values = overall.iloc[0]
        retained = [
            column for column in ["Weighted_Percent", "Unweighted_Percent", "N", "Missing_N"]
            if column in overall_values.index
        ]
        part = overall_values.loc[retained].to_frame(name="Overall")
        part.columns = pd.MultiIndex.from_tuples([("Overall", "All respondents")])
        matrix_parts.append(part)
        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            values = table[retained].T
            label = breakdown_labels.get(breakdown, breakdown)
            values.columns = pd.MultiIndex.from_tuples([
                (
                    label,
                    f"{_clean_group_label(column)} (n={int(table.loc[column, 'N'])})"
                    if "N" in table.columns and pd.notna(table.loc[column, "N"])
                    else _clean_group_label(column),
                )
                for column in values.columns
            ])
            matrix_parts.append(values)
    else:
        return pd.DataFrame()

    if not matrix_parts:
        return pd.DataFrame()
    matrix = pd.concat(matrix_parts, axis=1)
    matrix.index.name = "Statistic" if analysis_type in {"continuous", "indicator"} else "Response"
    return matrix


def write_consolidated_question(
    ws: Any,
    variable: str,
    row: int,
    *,
    analysis_type: str,
    qmeta: pd.DataFrame,
    descriptive_outputs: Mapping[str, Mapping[Any, pd.DataFrame]],
    report_results: pd.DataFrame,
    breakdowns: Sequence[str],
    breakdown_labels: Mapping[str, str],
    alpha: float,
    style: Mapping[str, str],
    title_override: str | None = None,
    detail_variables: Sequence[str] = (),
) -> int:
    """Write one question as a cohesive overall-plus-breakdown report table."""
    matrix = build_consolidated_question_matrix(
        variable,
        analysis_type,
        descriptive_outputs=descriptive_outputs,
        breakdowns=breakdowns,
        breakdown_labels=breakdown_labels,
    )
    if matrix.empty:
        return row

    label = get_question_label(qmeta, variable)
    heading = title_override or f"Question {question_number(variable)}. {label}"
    main_end_column = 1 + len(matrix.columns)
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=main_end_column
    )
    title = ws.cell(row=row, column=1, value=heading)
    title.font = Font(bold=True, size=12, color=style["primary_color"])
    title.alignment = Alignment(wrap_text=True)
    table_start_row = row + 1

    thin_grey = Side(style="thin", color=style["border_color"])
    first_header = ws.cell(
        row=table_start_row,
        column=1,
        value=matrix.index.name or "Response",
    )
    first_header.font = Font(bold=True, color=style["text_on_primary"])
    first_header.fill = PatternFill("solid", fgColor=style["primary_color"])
    first_header.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(
        start_row=table_start_row,
        start_column=1,
        end_row=table_start_row + 1,
        end_column=1,
    )

    column = 2
    for dimension in matrix.columns.get_level_values(0).unique():
        positions = [
            index for index, value in enumerate(matrix.columns.get_level_values(0))
            if value == dimension
        ]
        start, end = column, column + len(positions) - 1
        if end > start:
            ws.merge_cells(
                start_row=table_start_row,
                start_column=start,
                end_row=table_start_row,
                end_column=end,
            )
        cell = ws.cell(row=table_start_row, column=start, value=str(dimension))
        cell.font = Font(bold=True, color=style["text_on_primary"])
        cell.fill = PatternFill("solid", fgColor=style["primary_color"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for merged_column in range(start, end + 1):
            ws.cell(row=table_start_row, column=merged_column).fill = PatternFill(
                "solid", fgColor=style["primary_color"]
            )
        for offset, position in enumerate(positions):
            group_cell = ws.cell(
                row=table_start_row + 1,
                column=start + offset,
                value=str(matrix.columns[position][1]),
            )
            group_cell.font = Font(bold=True, color=style["text_on_primary"])
            group_cell.fill = PatternFill("solid", fgColor=style["secondary_color"])
            group_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        column = end + 1

    data_start = table_start_row + 2
    for row_offset, (index, values) in enumerate(matrix.iterrows()):
        excel_row = data_start + row_offset
        label_cell = ws.cell(row=excel_row, column=1, value=str(index))
        label_cell.alignment = Alignment(vertical="center", wrap_text=True)
        label_cell.border = Border(bottom=thin_grey)
        for column_offset, value in enumerate(values, 2):
            cell = ws.cell(
                row=excel_row,
                column=column_offset,
                value=None if pd.isna(value) else value,
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_grey)
            if isinstance(value, (float, np.floating)):
                cell.number_format = "0.0"

    main_end_row = data_start + len(matrix) - 1
    summary_row = main_end_row + 1
    for breakdown in breakdowns:
        selected = report_results.loc[
            (report_results["variable"] == variable)
            & (report_results["breakdown"] == breakdown)
        ]
        if selected.empty:
            continue
        comparison_label = breakdown_labels.get(breakdown, breakdown)
        summary_row = write_statistical_summary(
            ws,
            selected.iloc[0],
            summary_row,
            style,
            comparison_label,
            alpha,
            title=f"{comparison_label} comparison",
        )

    detail_end_row = row
    detail_column = main_end_column + 3
    for detail_variable in detail_variables:
        detail = descriptive_outputs["overall"].get(detail_variable)
        if not isinstance(detail, pd.DataFrame) or detail.empty:
            continue
        detail_title = (
            f"Follow-up / Other responses — "
            f"{get_question_label(qmeta, detail_variable)}"
        )
        detail_end_row = write_dataframe(
            ws,
            detail,
            row + 1,
            style,
            detail_title,
            start_column=detail_column,
        )
        detail_column += len(detail.columns) + 4

    return max(summary_row, detail_end_row, main_end_row + 1) + 2


def _write_stacked_header(
    ws: Any,
    row: int,
    columns: Sequence[object],
    style: Mapping[str, str],
) -> None:
    """Write one compact report-table header."""
    for column, value in enumerate(columns, 1):
        cell = ws.cell(row=row, column=column, value=str(value))
        cell.font = Font(bold=True, color=style["text_on_primary"])
        cell.fill = PatternFill("solid", fgColor=style["primary_color"])
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )


def _write_stacked_rows(
    ws: Any,
    row: int,
    table: pd.DataFrame,
    style: Mapping[str, str],
) -> int:
    """Write dataframe rows beneath an existing compact header."""
    thin_grey = Side(style="thin", color=style["border_color"])
    values = table.copy()
    for record_index, (index, record) in enumerate(values.iterrows(), row):
        is_total = str(index).startswith("Total")
        row_values = [index, *record.tolist()]
        for column, value in enumerate(row_values, 1):
            cell = ws.cell(
                row=record_index,
                column=column,
                value=None if pd.isna(value) else value,
            )
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=thin_grey)
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    "solid", fgColor=style["total_row_color"]
                )
            if isinstance(value, (float, np.floating)):
                cell.number_format = "0.0"
    return row + len(values)


def _write_breakdown_label(
    ws: Any,
    row: int,
    label: str,
    end_column: int,
    style: Mapping[str, str],
) -> int:
    """Write a full-width breakdown subheading."""
    if end_column > 1:
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=end_column,
        )
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color=style["primary_color"])
    cell.fill = PatternFill("solid", fgColor=style["total_row_color"])
    return row + 1


def write_stacked_question(
    ws: Any,
    variable: str,
    row: int,
    *,
    analysis_type: str,
    qmeta: pd.DataFrame,
    descriptive_outputs: Mapping[str, Mapping[Any, pd.DataFrame]],
    report_results: pd.DataFrame,
    breakdowns: Sequence[str],
    breakdown_labels: Mapping[str, str],
    alpha: float,
    style: Mapping[str, str],
    title_override: str | None = None,
    detail_variables: Sequence[str] = (),
) -> int:
    """Write a Word-friendly question table with stacked breakdown blocks."""
    overall = descriptive_outputs["overall"].get(variable)
    if not isinstance(overall, pd.DataFrame) or overall.empty:
        return row

    label = get_question_label(qmeta, variable)
    heading = title_override or question_heading(variable, label)
    title = ws.cell(row=row, column=1, value=heading)
    title.font = Font(bold=True, size=12, color=style["primary_color"])
    title.alignment = Alignment(wrap_text=True)
    row += 1

    if analysis_type in {"categorical_single", "open_text"}:
        response_columns = [str(value) for value in overall.index]
        value_column = next(
            (
                column
                for column in ["Weighted_Percent", "Percent", "Frequency"]
                if column in overall.columns
            ),
            None,
        )
        if value_column is None:
            return row
        include_total = value_column != "Frequency"
        columns = ["Analysis group", *response_columns]
        if include_total:
            columns.append("Total")
        _write_stacked_header(ws, row, columns, style)
        row += 1
        valid_n = (
            int(overall["Valid_N"].max())
            if "Valid_N" in overall.columns and overall["Valid_N"].notna().any()
            else int(overall["Frequency"].sum())
        )
        overall_values = overall[value_column].reindex(response_columns)
        overall_table = overall_values.to_frame().T
        overall_table.index = [f"Overall (n={valid_n})"]
        if include_total:
            overall_table["Total"] = 100.0
        row = _write_stacked_rows(ws, row, overall_table, style)

        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            breakdown_label = breakdown_labels.get(breakdown, breakdown)
            row = _write_breakdown_label(
                ws, row, breakdown_label, len(columns), style
            )
            values = table.reindex(columns=response_columns + ["Total"])
            row = _write_stacked_rows(ws, row, values, style)
            selected = report_results.loc[
                (report_results["variable"] == variable)
                & (report_results["breakdown"] == breakdown)
            ]
            if not selected.empty:
                row = write_statistical_summary(
                    ws,
                    selected.iloc[0],
                    row,
                    style,
                    breakdown_label,
                    alpha,
                    title="Statistical comparison",
                    end_column=len(columns),
                )

    elif analysis_type == "multiple_response":
        response_options = overall["Option"].astype(str).tolist()
        value_column = (
            "Weighted_Respondent_%"
            if "Weighted_Respondent_%" in overall.columns
            else "Respondent_%"
        )
        valid_n = int(overall["Valid_N"].max())
        overall_values = overall.set_index("Option")[value_column].reindex(
            response_options
        )
        overall_values = overall_values.to_frame()
        overall_values.columns = pd.MultiIndex.from_tuples([
            ("Overall", f"Weighted % (n={valid_n})")
        ])
        matrix_parts = [overall_values]
        comparison_summaries: list[tuple[str, str]] = []

        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            breakdown_label = breakdown_labels.get(breakdown, breakdown)
            reset = table.reset_index()
            pivoted = reset.pivot(
                index="Option", columns="Group", values=value_column
            ).reindex(response_options)
            bases = (
                reset.groupby("Group", sort=False)["Valid_N"].max()
                if "Valid_N" in reset.columns
                else pd.Series(dtype=float)
            )
            pivoted.columns = pd.MultiIndex.from_tuples([
                (
                    breakdown_label,
                    (
                        f"{group} (n={int(bases.loc[group])})"
                        if group in bases.index and pd.notna(bases.loc[group])
                        else str(group)
                    ),
                )
                for group in pivoted.columns
            ])
            matrix_parts.append(pivoted)

            selected = report_results.loc[
                report_results["variable"].astype(str).str.startswith(
                    f"{variable}::"
                )
                & report_results["breakdown"].eq(breakdown)
            ]
            if not selected.empty:
                completed = selected.loc[selected["status"].eq("completed")]
                significant = (
                    completed["p_value"].lt(alpha).any()
                    if not completed.empty
                    else False
                )
                summary = (
                    "At least one response option differed significantly "
                    f"across {breakdown_label}."
                    if significant
                    else "No statistically significant differences were detected "
                    f"for the eligible response options across {breakdown_label}."
                )
                comparison_summaries.append((breakdown_label, summary))

        matrix = pd.concat(matrix_parts, axis=1).reindex(response_options)
        table_end_column = 1 + len(matrix.columns)
        thin_grey = Side(style="thin", color=style["border_color"])

        response_header = ws.cell(row=row, column=1, value="Response")
        response_header.font = Font(bold=True, color=style["text_on_primary"])
        response_header.fill = PatternFill(
            "solid", fgColor=style["primary_color"]
        )
        response_header.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row + 1,
            end_column=1,
        )

        column = 2
        for dimension in matrix.columns.get_level_values(0).unique():
            positions = [
                index
                for index, value in enumerate(matrix.columns.get_level_values(0))
                if value == dimension
            ]
            start, end = column, column + len(positions) - 1
            if end > start:
                ws.merge_cells(
                    start_row=row,
                    start_column=start,
                    end_row=row,
                    end_column=end,
                )
            dimension_cell = ws.cell(
                row=row, column=start, value=str(dimension)
            )
            dimension_cell.font = Font(
                bold=True, color=style["text_on_primary"]
            )
            dimension_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            for current_column in range(start, end + 1):
                ws.cell(row=row, column=current_column).fill = PatternFill(
                    "solid", fgColor=style["primary_color"]
                )
            for offset, position in enumerate(positions):
                group_cell = ws.cell(
                    row=row + 1,
                    column=start + offset,
                    value=str(matrix.columns[position][1]),
                )
                group_cell.font = Font(
                    bold=True, color=style["text_on_primary"]
                )
                group_cell.fill = PatternFill(
                    "solid", fgColor=style["secondary_color"]
                )
                group_cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            column = end + 1

        row += 2
        for response, values in matrix.iterrows():
            label_cell = ws.cell(row=row, column=1, value=str(response))
            label_cell.alignment = Alignment(vertical="center", wrap_text=True)
            label_cell.border = Border(bottom=thin_grey)
            for column, value in enumerate(values, 2):
                cell = ws.cell(
                    row=row,
                    column=column,
                    value=None if pd.isna(value) else value,
                )
                cell.number_format = "0.0"
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                cell.border = Border(bottom=thin_grey)
            row += 1

        for breakdown_label, summary in comparison_summaries:
            ws.cell(
                row=row,
                column=1,
                value=f"{breakdown_label} comparison",
            ).font = Font(bold=True, color=style["primary_color"])
            if table_end_column > 1:
                ws.merge_cells(
                    start_row=row,
                    start_column=2,
                    end_row=row,
                    end_column=table_end_column,
                )
            summary_cell = ws.cell(row=row, column=2, value=summary)
            summary_cell.alignment = Alignment(wrap_text=True)
            summary_cell.fill = PatternFill(
                "solid", fgColor=style["summary_fill_color"]
            )
            row += 1

    elif analysis_type == "continuous":
        statistics = [
            column
            for column in ["N", "Mean", "Median", "SD", "IQR", "Min", "Max", "Missing"]
            if column in overall.columns
        ]
        columns = ["Analysis group", *statistics]
        _write_stacked_header(ws, row, columns, style)
        row += 1
        overall_table = overall[statistics].copy()
        overall_table.index = ["Overall"]
        row = _write_stacked_rows(ws, row, overall_table, style)
        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            breakdown_label = breakdown_labels.get(breakdown, breakdown)
            row = _write_breakdown_label(
                ws, row, breakdown_label, len(columns), style
            )
            row = _write_stacked_rows(
                ws, row, table.reindex(columns=statistics), style
            )

    elif analysis_type == "indicator":
        statistics = [
            column
            for column in ["Weighted_Percent", "Unweighted_Percent", "N", "Missing_N"]
            if column in overall.columns
        ]
        columns = ["Analysis group", *statistics]
        _write_stacked_header(ws, row, columns, style)
        row += 1
        overall_table = overall[statistics].copy()
        overall_table.index = ["Overall"]
        row = _write_stacked_rows(ws, row, overall_table, style)
        for breakdown in breakdowns:
            table = descriptive_outputs["breakdowns"].get((variable, breakdown))
            if not isinstance(table, pd.DataFrame) or table.empty:
                continue
            breakdown_label = breakdown_labels.get(breakdown, breakdown)
            row = _write_breakdown_label(
                ws, row, breakdown_label, len(columns), style
            )
            row = _write_stacked_rows(
                ws, row, table.reindex(columns=statistics), style
            )
            selected = report_results.loc[
                (report_results["variable"] == variable)
                & (report_results["breakdown"] == breakdown)
            ]
            if not selected.empty:
                row = write_statistical_summary(
                    ws,
                    selected.iloc[0],
                    row,
                    style,
                    breakdown_label,
                    alpha,
                    title="Statistical comparison",
                    end_column=len(columns),
                )
    else:
        return row

    for detail_variable in detail_variables:
        detail = descriptive_outputs["overall"].get(detail_variable)
        if not isinstance(detail, pd.DataFrame) or detail.empty:
            continue
        row += 1
        row = write_dataframe(
            ws,
            detail,
            row,
            style,
            (
                "Follow-up / Other responses — "
                f"{get_question_label(qmeta, detail_variable)}"
            ),
        )
    return row + 2


def build_statistical_interpretation(result: Mapping[str, Any], comparison_label: str, alpha: float) -> str | None:
    """Build the standard report sentence for one inferential result."""
    if result.get("status") == "suppressed" or not result.get("test"):
        reason = result.get("suppression_reason") or result.get("interpretation")
        return f"Statistical comparison not performed: {reason}"
    p_value = result.get("p_value")
    if pd.isna(p_value):
        return None
    adjusted_p = result.get("adjusted_p_value")
    decision_p = adjusted_p if pd.notna(adjusted_p) else p_value
    finding = (
        "Statistically significant differences were observed"
        if decision_p < alpha
        else "No statistically significant differences were detected"
    )
    p_text = "p < 0.001" if p_value < 0.001 else f"p = {p_value:.3f}"
    if pd.notna(adjusted_p):
        adjusted_text = (
            "adjusted p < 0.001"
            if adjusted_p < 0.001
            else f"adjusted p = {adjusted_p:.3f}"
        )
        p_text = f"{p_text}; {adjusted_text}"
    test_name = str(result["test"]).replace("Rao-Scott", "Rao–Scott")
    return f"{finding} across {comparison_label} ({test_name}, {p_text})."


def write_statistical_summary(
    ws: Any,
    result: Mapping[str, Any],
    row: int,
    style: Mapping[str, str],
    comparison_label: str,
    alpha: float,
    title: str = "Statistical summary",
    end_column: int = 8,
) -> int:
    """Write a formatted statistical-summary row when a result is reportable."""
    text = build_statistical_interpretation(result, comparison_label, alpha)
    if not text:
        return row
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color=style["primary_color"])
    ws.merge_cells(
        start_row=row,
        start_column=2,
        end_row=row,
        end_column=max(2, end_column),
    )
    summary = ws.cell(row=row, column=2, value=text)
    summary.alignment = Alignment(wrap_text=True, vertical="center")
    summary.fill = PatternFill("solid", fgColor=style["summary_fill_color"])
    return row + 1


def write_question(
    ws: Any,
    variable: str,
    row: int,
    *,
    qmeta: pd.DataFrame,
    descriptive_outputs: Mapping[str, Mapping[Any, pd.DataFrame]],
    report_results: pd.DataFrame,
    breakdowns: Sequence[str],
    breakdown_labels: Mapping[str, str],
    alpha: float,
    style: Mapping[str, str],
    title_override: str | None = None,
) -> int:
    """Write one report block from precomputed outputs and explicit metadata."""
    label = get_question_label(qmeta, variable)
    heading = title_override or f"Question {question_number(variable)}. {label}"
    title = ws.cell(row=row, column=1, value=heading)
    title.font = Font(bold=True, size=12, color=style["primary_color"])
    title.alignment = Alignment(wrap_text=True)
    row += 1
    overall = descriptive_outputs["overall"].get(variable)
    if isinstance(overall, pd.DataFrame) and not overall.empty:
        row = write_dataframe(ws, overall, row, style, "Overall")
    for breakdown in breakdowns:
        table = descriptive_outputs["breakdowns"].get((variable, breakdown))
        if isinstance(table, pd.DataFrame) and not table.empty:
            label = breakdown_labels.get(breakdown, get_question_label(qmeta, breakdown))
            row = write_dataframe(ws, table, row, style, label)
            selected = report_results.loc[(report_results["variable"] == variable) & (report_results["breakdown"] == breakdown)]
            if not selected.empty:
                row = write_statistical_summary(ws, selected.iloc[0], row, style, label, alpha)
            row += 1
    return row + 1


def write_section_heading(ws: Any, text: str, row: int, style: Mapping[str, str], size: int = 14) -> None:
    """Write a plain styled section heading in a worksheet."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=size, color=style["primary_color"])


def format_sheet(ws: Any, max_width: int = 38, preview_rows: int = 250) -> None:
    """Apply generic worksheet view and column-width formatting."""
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for column in range(1, ws.max_column + 1):
        values = [len(str(ws.cell(row, column).value or "")) for row in range(1, min(ws.max_row, preview_rows) + 1)]
        ws.column_dimensions[get_column_letter(column)].width = min(max(max(values, default=10) + 2, 12), max_width)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 28)


def write_question_analysis_workbook(
    analysis: pd.DataFrame,
    output_path: str,
    *,
    information_rows: Sequence[tuple[str, str]],
    style: Mapping[str, str],
    data_sheet_name: str = "Question_Analysis",
) -> None:
    """Write static, pre-calculated survey-analysis results and guidance."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        analysis.to_excel(writer, sheet_name=data_sheet_name, index=False)
        information = pd.DataFrame(information_rows, columns=["Topic", "Guidance"])
        information.to_excel(writer, sheet_name="Information", index=False)

        header_fill = PatternFill("solid", fgColor=style["primary_color"])
        for sheet_name in (data_sheet_name, "Information"):
            worksheet = writer.sheets[sheet_name]
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color=style["text_on_primary"])
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            worksheet.auto_filter.ref = worksheet.dimensions
            format_sheet(worksheet)

        information_ws = writer.sheets["Information"]
        information_ws.column_dimensions["A"].width = 34
        information_ws.column_dimensions["B"].width = 120
        for row in information_ws.iter_rows(min_row=2, max_col=2):
            row[0].font = Font(bold=True)
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_number in range(2, information_ws.max_row + 1):
            information_ws.row_dimensions[row_number].height = 42
